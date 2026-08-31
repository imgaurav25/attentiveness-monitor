"""
excel_logger.py
-----------------
Writes two kinds of records to one workbook, on two separate sheets so each
has clean, relevant columns:

  Sheet "Attentiveness Log"
    One row per inattentiveness lapse (Looking away / Eyes closed / Away
    from camera), with the REAL original start time and REAL total
    duration, and the student's identity when recognized (known mode).

  Sheet "Unknown Person Log"
    One row per unknown-person alert (known/classroom mode). Includes the
    NEARBY registered student (if one was in frame at the time -- found by
    face-position proximity, never assumed), the unknown person's own
    tracking ID, how long they were present, and paths to both the
    unknown person's face crop and the full camera frame.
"""

import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

ATTENTIVENESS_SHEET = "Attentiveness Log"
UNKNOWN_SHEET = "Unknown Person Log"

ATTENTIVENESS_HEADERS = [
    "Event Timestamp (start)", "Date", "Time",
    "Member ID", "Name", "Roll No", "Class", "Department", "Year",
    "Event Type", "Duration (s)", "Snapshot",
]

MEETING_ALERT_SHEET = "Meeting Alert Log"
MEETING_ALERT_HEADERS = [
    "Timestamp", "Participant ID", "Name", "Roll No", "Class", "Department", "Year",
    "Event Type", "Duration (s)", "Message", "Face Snapshot", "Full Frame Snapshot",
]

UNKNOWN_HEADERS = [
    "Timestamp", "Date", "Time",
    "Registered Student", "Roll No", "Class", "Department", "Year",
    "Unknown ID", "Event Type", "Duration (s)",
    "Unknown Face Snapshot", "Full Frame Snapshot",
]


class ExcelLogger:
    def __init__(self, path):
        self.path = path
        self._lock = threading.Lock()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if not os.path.exists(path):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = ATTENTIVENESS_SHEET
            ws1.append(ATTENTIVENESS_HEADERS)
            self._autosize(ws1, ATTENTIVENESS_HEADERS)

            ws2 = wb.create_sheet(UNKNOWN_SHEET)
            ws2.append(UNKNOWN_HEADERS)
            self._autosize(ws2, UNKNOWN_HEADERS)

            wb.save(path)
        else:
            # Backfill the unknown-person sheet for workbooks created by an
            # older version of this app that only had the attentiveness sheet.
            wb = load_workbook(path)
            changed = False
            if UNKNOWN_SHEET not in wb.sheetnames:
                ws2 = wb.create_sheet(UNKNOWN_SHEET)
                ws2.append(UNKNOWN_HEADERS)
                self._autosize(ws2, UNKNOWN_HEADERS)
                changed = True
            if MEETING_ALERT_SHEET not in wb.sheetnames:
                ws3 = wb.create_sheet(MEETING_ALERT_SHEET)
                ws3.append(MEETING_ALERT_HEADERS)
                self._autosize(ws3, MEETING_ALERT_HEADERS)
                changed = True
            if changed:
                wb.save(path)

    @staticmethod
    def _autosize(ws, headers):
        for col, header in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(14, len(header) + 4)

    @staticmethod
    def _identity_fields(identity):
        if not identity:
            return ("", "", "", "", "")
        return (
            identity.get("name", ""),
            identity.get("roll_no", ""),
            identity.get("class", ""),
            identity.get("department", ""),
            identity.get("year", ""),
        )

    # ------------------------------------------------------ attentiveness
    def log_event(self, member_id, reason, start_ts, duration_seconds, identity=None, snapshot_path=""):
        with self._lock:
            wb = load_workbook(self.path)
            ws = wb[ATTENTIVENESS_SHEET]
            dt = datetime.fromtimestamp(start_ts)
            name, roll_no, klass, dept, year = self._identity_fields(identity)
            ws.append([
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                f"Member {member_id}",
                name, roll_no, klass, dept, year,
                reason,
                round(duration_seconds, 1),
                os.path.basename(snapshot_path) if snapshot_path else "",
            ])
            wb.save(self.path)


    def log_meeting_alert(self, participant_id, name, event_type, timestamp, duration_seconds,
                          identity=None, message="", face_snapshot="", full_snapshot=""):
        with self._lock:
            wb = load_workbook(self.path)
            if MEETING_ALERT_SHEET not in wb.sheetnames:
                ws = wb.create_sheet(MEETING_ALERT_SHEET)
                ws.append(MEETING_ALERT_HEADERS)
            ws = wb[MEETING_ALERT_SHEET]
            dt = datetime.fromtimestamp(timestamp)
            roll_no = identity.get("roll_no", "") if identity else ""
            klass = identity.get("class", "") if identity else ""
            dept = identity.get("department", "") if identity else ""
            year = identity.get("year", "") if identity else ""
            ws.append([dt.strftime("%Y-%m-%d %H:%M:%S"), participant_id, name, roll_no, klass, dept, year,
                       event_type, round(float(duration_seconds), 1), message,
                       os.path.basename(face_snapshot) if face_snapshot else "",
                       os.path.basename(full_snapshot) if full_snapshot else ""])
            wb.save(self.path)

    # ------------------------------------------------------------ unknown
    def log_unknown_with_association(self, unknown_member_id, duration_seconds,
                                      nearby_identity=None, unknown_face_path="", full_frame_path="",
                                      timestamp=None, message=None):
        """nearby_identity: metadata dict of the closest REGISTERED student in
        frame at the time of the alert, or None if no registered student was
        present (never guessed/assumed -- see association.py).
        message: the human-readable event description -- "Unknown" when no
        registered student was nearby, or "Unknown present with <name>" when
        one was. Falls back to a generic label if not supplied."""
        with self._lock:
            wb = load_workbook(self.path)
            ws = wb[UNKNOWN_SHEET]
            dt = datetime.fromtimestamp(timestamp) if timestamp else datetime.now()
            name, roll_no, klass, dept, year = self._identity_fields(nearby_identity)
            event_type = message or ("Unknown present with " + name if name else "Unknown")
            ws.append([
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                name, roll_no, klass, dept, year,
                f"UNKNOWN-{unknown_member_id:02d}" if isinstance(unknown_member_id, int) else f"UNKNOWN-{unknown_member_id}",
                event_type,
                round(duration_seconds, 1),
                os.path.basename(unknown_face_path) if unknown_face_path else "",
                os.path.basename(full_frame_path) if full_frame_path else "",
            ])
            wb.save(self.path)

    # ------------------------------------------------------------ counts
    def count_events(self):
        """Total across both sheets (kept for backward compatibility)."""
        return self.count_attentiveness_events() + self.count_unknown_events()

    def count_attentiveness_events(self):
        with self._lock:
            wb = load_workbook(self.path, read_only=True)
            ws = wb[ATTENTIVENESS_SHEET]
            return max(0, ws.max_row - 1)

    def count_unknown_events(self):
        with self._lock:
            wb = load_workbook(self.path, read_only=True)
            ws = wb[UNKNOWN_SHEET]
            return max(0, ws.max_row - 1)
